from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

data = [
    (1, "ファイルを読む", 80),
    (2, "ファイルを作る", 85),
    (3, "フォルダ構造", 90),
    (4, "コピー管理", 85),
    (5, "タスク管理", 88),
    (6, "ファイル編集", 92),
    (7, "Grep検索", 90),
    (8, "Git基本", 88),
    (9, "デバッグ", 85),
    (10, "Agentツール", 95),
]

wb = Workbook()
ws = wb.active
ws.title = "学習記録"

headers = ["Day", "内容", "スコア"]
ws.append(headers)

header_font = Font(bold=True, color="FFFFFF", name="Arial")
header_fill = PatternFill("solid", start_color="1F4E79")

for col in range(1, 4):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for row in data:
    ws.append(row)

last_row = len(data) + 2
ws.cell(row=last_row, column=2).value = "平均スコア"
ws.cell(row=last_row, column=2).font = Font(bold=True, name="Arial")
ws.cell(row=last_row, column=3).value = f"=AVERAGE(C2:C{last_row - 1})"
ws.cell(row=last_row, column=3).font = Font(bold=True, name="Arial")

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 12

output_path = "day34/learning_report.xlsx"
wb.save(output_path)
print(f"Excelファイルを作成しました：{output_path}")
